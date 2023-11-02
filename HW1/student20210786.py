#!/usr/bin/python3

from openpyxl import load_workbook
from operator import itemgetter
import math


def get_num(stu_num) :
    A_num = stu_num * 0.3 # 22.2 
    APlus_num = A_num / 2 # 11.1
    A_num -= APlus_num  
    A_num = math.trunc(A_num) # 11
    APlus_num = math.trunc(APlus_num) # 11
    
    B_num = stu_num * 0.4 # 29.6 
    BPlus_num = B_num / 2 # 14.8   
    B_num -= BPlus_num # 14.8    
    B_num = math.trunc(B_num) # 14
    BPlus_num = math.trunc(BPlus_num) # 14
      
    C_num = stu_num - (A_num + B_num + APlus_num + BPlus_num) # 74 - 50 == 24
    CPlus_num = C_num // 2 # 12
    C_num -= CPlus_num # 12

    return A_num, APlus_num, B_num, BPlus_num, C_num, CPlus_num


wb = load_workbook(filename= 'student.xlsx')
sheet = wb['Sheet1']

info_list = []
stu_num = sheet.max_row - 1
A_num, APlus_num, B_num, BPlus_num, C_num, CPlus_num = get_num(stu_num)

for rowNum in range(2, stu_num + 2):
    info_dict = dict()
    total = 0
    id = sheet.cell(row = rowNum, column = 1).value
    midterm = sheet.cell(row = rowNum, column = 3).value * 0.3
    final = sheet.cell(row = rowNum, column = 4).value * 0.35
    hw = sheet.cell(row = rowNum, column = 5).value * 0.34
    attendance = sheet.cell(row = rowNum, column = 6).value * 0.01
    total = midterm + final + hw + attendance
    total = round(total, 2)
    sheet.cell(row = rowNum, column = 7, value = total)
    info_dict = {"id" : id, "total" : total, "rowNum" : rowNum}
    info_list.append(info_dict)

info_list = sorted(info_list, key = itemgetter("total"), reverse = True)

for i in range(0, APlus_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'A+')


for i in range(APlus_num, APlus_num + A_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'A')

for i in range(APlus_num + A_num, APlus_num + A_num + BPlus_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'B+')

for i in range(APlus_num + A_num + BPlus_num, APlus_num + A_num + BPlus_num + B_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'B')

for i in range(APlus_num + A_num + BPlus_num + B_num, APlus_num + A_num + BPlus_num + B_num + CPlus_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'C+')

for i in range(APlus_num + A_num + BPlus_num + B_num + CPlus_num, APlus_num + A_num + BPlus_num + B_num + CPlus_num + C_num) :
    if(info_list[i]['total'] < 40) :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'F')
    else :
        sheet.cell(row = info_list[i]['rowNum'], column = 8, value = 'C')

wb.save(filename = 'student.xlsx')
